import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue
ROW_MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light Green
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")       # Light Gray

# Border style
THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

class ExcelComparator:
    def __init__(self, file1_path, file2_path, sheet1_name=None, sheet2_name=None, 
                 highlight_missing=True, highlight_cell_diffs=True, 
                 highlight_row_matches=True, create_num_table=True):
        """
        Initialize the ExcelComparator with file paths and options.
        
        Args:
            file1_path (str): Path to first Excel file
            file2_path (str): Path to second Excel file
            sheet1_name (str, optional): Sheet name for first file. Defaults to first sheet.
            sheet2_name (str, optional): Sheet name for second file. Defaults to first sheet.
            highlight_missing (bool): Whether to highlight missing columns. Default True.
            highlight_cell_diffs (bool): Whether to highlight cell differences. Default True.
            highlight_row_matches (bool): Whether to highlight row matches/mismatches. Default True.
            create_num_table (bool): Whether to create numerical differences table. Default True.
        """
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name
        self.highlight_missing = highlight_missing
        self.highlight_cell_diffs = highlight_cell_diffs
        self.highlight_row_matches = highlight_row_matches
        self.create_num_table = create_num_table
        self.df1 = None
        self.df2 = None
        
        # If sheet names not provided, get first sheet name
        if not self.sheet1_name:
            wb = load_workbook(file1_path, read_only=True)
            self.sheet1_name = wb.sheetnames[0]
            wb.close()
            
        if not self.sheet2_name:
            wb = load_workbook(file2_path, read_only=True)
            self.sheet2_name = wb.sheetnames[0]
            wb.close()
    
    def are_equal(self, a, b):
        """Check if two values are equal, handling NaN cases"""
        if pd.isna(a) and pd.isna(b):
            return True
        if pd.isna(a) or pd.isna(b):
            return False
        return a == b
    
    def is_date(self, value):
        """Check if a value is a date"""
        return isinstance(value, (datetime, pd.Timestamp))
    
    def get_string_columns(self, df):
        """Get all string columns that are not dates"""
        string_cols = []
        for col in df.columns:
            # Skip date columns
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                continue
            
            # Include string columns and object columns that are likely strings
            if pd.api.types.is_string_dtype(df[col]) or \
               (pd.api.types.is_object_dtype(df[col]) and 
                all(isinstance(x, str) or pd.isna(x) for x in df[col].head(100))):
                string_cols.append(col)
        return string_cols
    
    def create_side_by_side_sheet(self, df1, df2, output_wb):
        """Create side-by-side comparison sheet with row matching column"""
        # Create sheet
        ws = output_wb.create_sheet("Side by Side Comparison")
        
        # Get common columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        
        # Get all string columns (excluding dates)
        str_cols1 = self.get_string_columns(df1)
        str_cols2 = self.get_string_columns(df2)
        all_str_cols = list(set(str_cols1) | set(str_cols2))
        
        # Create concatenation keys using all string columns
        concat_keys1 = {}
        concat_keys2 = {}
        
        # Create keys for df1
        for idx, row in df1.iterrows():
            key_parts = []
            for col in all_str_cols:
                if col in df1.columns:
                    val = row[col]
                    if pd.notna(val) and not self.is_date(val):
                        key_parts.append(str(val))
            concat_keys1[idx] = "_".join(key_parts) if key_parts else None
        
        # Create keys for df2
        for idx, row in df2.iterrows():
            key_parts = []
            for col in all_str_cols:
                if col in df2.columns:
                    val = row[col]
                    if pd.notna(val) and not self.is_date(val):
                        key_parts.append(str(val))
            concat_keys2[idx] = "_".join(key_parts) if key_parts else None
        
        # Create sets of keys for matching
        keys1_set = set(concat_keys1.values())
        keys2_set = set(concat_keys2.values())
        
        # Create match status columns in original dataframes
        df1['Match Status'] = "Not Matched"
        df2['Match Status'] = "Not Matched"
        
        # Mark matched rows in both dataframes
        for key in keys1_set:
            if key in keys2_set:
                # Mark all rows with this key in both files as matched
                df1.loc[df1.index.isin([idx for idx, k in concat_keys1.items() if k == key]), 'Match Status'] = "Matched"
                df2.loc[df2.index.isin([idx for idx, k in concat_keys2.items() if k == key]), 'Match Status'] = "Matched"
        
        # Write headers
        header_row = list(df1.columns) + list(df2.columns)
        ws.append(header_row)
        
        # Apply header styling
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        
        # Write data row by row
        max_rows = max(len(df1), len(df2))
        for i in range(max_rows):
            row_data = []
            
            # Add File1 data if exists
            if i < len(df1):
                row_data.extend(df1.iloc[i].tolist())
            else:
                row_data.extend([""] * len(df1.columns))
            
            # Add File2 data if exists
            if i < len(df2):
                row_data.extend(df2.iloc[i].tolist())
            else:
                row_data.extend([""] * len(df2.columns))
            
            ws.append(row_data)
            
            # Apply row matching highlighting
            if self.highlight_row_matches:
                file1_match = df1.iloc[i]['Match Status'] if i < len(df1) else None
                file2_match = df2.iloc[i]['Match Status'] if i < len(df2) else None
                
                # Calculate column positions
                file1_start_col = 1
                file1_end_col = len(df1.columns)
                file2_start_col = file1_end_col + 1
                file2_end_col = file2_start_col + len(df2.columns) - 1
                match_status_col = file2_end_col + 1
                
                # Apply styling
                if file1_match == "Matched":
                    for col_idx in range(file1_start_col, file1_end_col + 1):
                        ws.cell(row=i+2, column=col_idx).fill = ROW_MATCH_FILL
                else:
                    for col_idx in range(file1_start_col, file1_end_col + 1):
                        ws.cell(row=i+2, column=col_idx).fill = ROW_MISSING_FILL
                
                if file2_match == "Matched":
                    for col_idx in range(file2_start_col, file2_end_col + 1):
                        ws.cell(row=i+2, column=col_idx).fill = ROW_MATCH_FILL
                else:
                    for col_idx in range(file2_start_col, file2_end_col + 1):
                        ws.cell(row=i+2, column=col_idx).fill = ROW_MISSING_FILL
            
            # Apply cell difference highlighting
            if self.highlight_cell_diffs:
                if i < len(df1) and i < len(df2):
                    for col in common_cols:
                        val1 = df1.at[i, col]
                        val2 = df2.at[i, col]
                        
                        if not self.are_equal(val1, val2):
                            col_idx1 = list(df1.columns).index(col) + 1
                            col_idx2 = list(df2.columns).index(col) + len(df1.columns) + 1
                            
                            ws.cell(row=i+2, column=col_idx1).fill = CELL_DIFF_FILL
                            ws.cell(row=i+2, column=col_idx2).fill = CELL_DIFF_FILL
            
            # Apply borders
            for col_idx in range(1, len(header_row) + 1):
                ws.cell(row=i+2, column=col_idx).border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, len(header_row) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = "A2"
        
        return len(df1[df1['Match Status'] == "Matched"]), len(df1[df1['Match Status'] == "Not Matched"]), len(df2[df2['Match Status'] == "Not Matched"])
    
    def compare_headers(self, df1, df2, output_wb):
        """Compare and highlight header differences"""
        headers1 = set(df1.columns)
        headers2 = set(df2.columns)
        common = headers1 & headers2
        unique1 = headers1 - headers2
        unique2 = headers2 - headers1
        
        # Create header comparison sheet
        ws = output_wb.create_sheet("Header Comparison")
        ws.append(["Header", "Status", "File 1 Presence", "File 2 Presence"])
        
        # Add common headers
        for header in sorted(common):
            ws.append([header, "Common", "✓", "✓"])
        
        # Add unique headers if option is enabled
        if self.highlight_missing:
            for header in sorted(unique1):
                ws.append([header, "Unique to File 1", "✓", ""])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
            
            for header in sorted(unique2):
                ws.append([header, "Unique to File 2", "", "✓"])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
        
        # Apply formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(bold=(cell.row == 1))
                cell.border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, 5):  # We have 4 columns in this sheet
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def analyze_row_matches(self, df1, df2, output_wb, matched_count, unmatched1_count, unmatched2_count):
        """Analyze and highlight row matches between files"""
        # Create row matching sheet
        ws = output_wb.create_sheet("Row Matching Analysis")
        
        # Summary section
        ws.append(["Row Matching Summary"])
        ws.append(["", ""])
        ws.append(["Total Rows in File1", len(df1)])
        ws.append(["Total Rows in File2", len(df2)])
        ws.append(["Matched Rows", matched_count])
        ws.append(["Unmatched Rows in File1", unmatched1_count])
        ws.append(["Unmatched Rows in File2", unmatched2_count])
        ws.append([""])
        
        # Key generation method
        ws.append(["Key Generation Method:"])
        ws.append(["Automatically concatenated all non-date string columns"])
        ws.append([""])
        ws.append(["Note: Rows are considered matched if any row in File1 matches any row in File2"])
        
        # Apply styling to summary
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)
        
        for row in ws.iter_rows(min_row=3, max_row=7):
            for cell in row:
                cell.font = Font(bold=(cell.column == 1))
        
        # Auto-size columns
        for col_idx in range(1, 3):  # We have 2 columns in this sheet
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def compare_numeric_values(self, df1, df2, output_wb):
        """Create numerical comparison table for common numeric columns"""
        # Identify common numeric columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        num_cols = [col for col in common_cols 
                    if pd.api.types.is_numeric_dtype(df1[col]) and 
                    pd.api.types.is_numeric_dtype(df2[col])]
        
        if not num_cols:
            return
        
        # Create sheet
        ws = output_wb.create_sheet("Numeric Comparison")
        headers = ["Column", "Row", "File1 Value", "File2 Value", "Absolute Diff", "Relative Diff"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        # Compare values
        for col in num_cols:
            for i in range(min(len(df1), len(df2))):
                val1 = df1[col].iloc[i]
                val2 = df2[col].iloc[i]
                
                if pd.isna(val1) or pd.isna(val2) or val1 == val2:
                    continue
                    
                abs_diff = abs(val1 - val2)
                rel_diff = abs_diff / max(abs(val1), abs(val2)) if max(abs(val1), abs(val2)) != 0 else float('inf')
                
                ws.append([col, i+1, val1, val2, abs_diff, rel_diff])
                
                # Highlight significant differences (>10%)
                if rel_diff > 0.1:
                    for col_idx in range(1, 7):
                        cell = ws.cell(ws.max_row, col_idx)
                        cell.fill = NUM_DIFF_FILL
                
                # Apply borders
                for col_idx in range(1, 7):
                    cell = ws.cell(ws.max_row, col_idx)
                    cell.border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, 7):  # We have 6 columns in this sheet
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def compare(self, output_file=None):
        """
        Compare the Excel files and save results.
        
        Args:
            output_file (str, optional): Path to save comparison results. If None, returns the workbook object.
        
        Returns:
            Workbook: If output_file is None, returns the workbook object
            str: If output_file is provided, returns the path to saved file
        """
        try:
            # Read data
            df1 = pd.read_excel(self.file1_path, sheet_name=self.sheet1_name)
            df2 = pd.read_excel(self.file2_path, sheet_name=self.sheet2_name)
            
            # Create comparison workbook
            output_wb = Workbook()
            output_wb.remove(output_wb.active)
            
            # 1. Compare headers
            self.compare_headers(df1, df2, output_wb)
            
            # 2. Create side-by-side comparison sheet
            matched_count, unmatched1_count, unmatched2_count = self.create_side_by_side_sheet(df1, df2, output_wb)
            
            # 3. Row matching analysis
            self.analyze_row_matches(df1, df2, output_wb, matched_count, unmatched1_count, unmatched2_count)
            
            # 4. Numerical differences
            if self.create_num_table:
                self.compare_numeric_values(df1, df2, output_wb)
            
            if output_file:
                output_wb.save(output_file)
                return output_file
            else:
                return output_wb
                
        except Exception as e:
            raise Exception(f"An error occurred during comparison: {str(e)}")

# Example usage:
if __name__ == "__main__":
    # Initialize comparator with file paths and options
    comparator = ExcelComparator(
        file1_path="file1.xlsx",
        file2_path="file2.xlsx",
        sheet1_name="Sheet1",  # optional
        sheet2_name="Sheet1",  # optional
        highlight_missing=True,
        highlight_cell_diffs=True,
        highlight_row_matches=True,
        create_num_table=True
    )
    
    # Run comparison and save results
    try:
        result_path = comparator.compare(output_file="comparison_results.xlsx")
        print(f"Comparison saved successfully to: {result_path}")
    except Exception as e:
        print(f"Error during comparison: {e}")
