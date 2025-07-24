import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Optional, Union

# --- Define Highlighting Styles ---
CELL_DIFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
ROW_MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
ROW_MISMATCH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light Gray
HEADER_FILL = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid") # Dark Gray
SUMMARY_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
TOTAL_ROW_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light Blue
BLACK_SEPARATOR_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

# --- Border Style ---
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

class ExcelComparator:
    """
    A class to compare two Excel sheets, offering both key-based and index-based comparison.
    
    Generates a detailed report highlighting differences, missing rows, and a summary.
    """
    def __init__(self, 
                 file1_path: str, 
                 file2_path: str, 
                 sheet1_name: Optional[str] = None, 
                 sheet2_name: Optional[str] = None,
                 key_columns: Optional[List[str]] = None,
                 total_row_identifier: str = "Total"):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.key_columns = key_columns if key_columns else []
        self.total_row_identifier = total_row_identifier

        # Automatically determine sheet names if not provided
        self.sheet1_name = sheet1_name or self._get_first_sheet_name(file1_path)
        self.sheet2_name = sheet2_name or self._get_first_sheet_name(file2_path)

    def _get_first_sheet_name(self, file_path: str) -> str:
        """Gets the name of the first sheet in an Excel workbook."""
        wb = load_workbook(file_path, read_only=True)
        return wb.sheetnames[0]

    def _are_equal(self, a, b) -> bool:
        """Robustly checks if two values are equal, handling NaN, datatypes, and float tolerance."""
        if pd.isna(a) and pd.isna(b):
            return True
        if pd.isna(a) or pd.isna(b):
            return False
        if isinstance(a, (datetime, pd.Timestamp)) and isinstance(b, (datetime, pd.Timestamp)):
            return a == b
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            return abs(a - b) < 1e-9 # Tolerance for floating point differences
        return str(a).strip() == str(b).strip()

    def _filter_total_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filters out rows that are identified as 'total' rows."""
        if self.total_row_identifier and not df.empty and df.shape[1] > 0:
            first_col = df.columns[0]
            return df[~df[first_col].astype(str).str.contains(self.total_row_identifier, case=False, na=False)]
        return df

    def compare(self, output_file: str = "comparison_results.xlsx") -> str:
        """
        Main method to perform the comparison and generate the output Excel file.
        
        Returns:
            The path to the generated output file.
        """
        try:
            df1 = pd.read_excel(self.file1_path, sheet_name=self.sheet1_name)
            df2 = pd.read_excel(self.file2_path, sheet_name=self.sheet2_name)

            # --- Data Preparation ---
            # Create dataframes for summation, excluding any pre-existing "Total" rows
            df1_sum_data = self._filter_total_rows(df1)
            df2_sum_data = self._filter_total_rows(df2)

            # Use index for comparison if no key columns are specified
            use_index_key = not self.key_columns
            if use_index_key:
                self.key_columns = ['__temp_index_key__']
                df1[self.key_columns[0]] = df1.index
                df2[self.key_columns[0]] = df2.index

            # --- Merging Data ---
            comparison_df = pd.merge(df1, df2, on=self.key_columns, how='outer', suffixes=('_f1', '_f2'), indicator=True)
            
            # --- Workbook Creation ---
            output_wb = Workbook()
            output_wb.remove(output_wb.active) # Remove default sheet

            # --- Generate Report Sheets ---
            self._generate_summary_sheet(output_wb, comparison_df, df1.shape[0], df2.shape[0])
            self._generate_comparison_sheet(output_wb, comparison_df, df1_sum_data, df2_sum_data, use_index_key)

            output_wb.save(output_file)
            return output_file
            
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Error: Input file not found - {e.filename}")
        except Exception as e:
            raise Exception(f"An unexpected error occurred during comparison: {str(e)}")

    def _generate_summary_sheet(self, wb: Workbook, comparison_df: pd.DataFrame, total_rows1: int, total_rows2: int):
        """Creates the summary sheet with high-level statistics."""
        ws = wb.create_sheet("Summary", 0)
        
        # --- Data Calculation ---
        matched_keys = comparison_df['_merge'] == 'both'
        mismatched_rows = comparison_df[matched_keys]['Status'] == 'Mismatch'
        
        summary_data = {
            'File 1 Details': {'File Path': self.file1_path, 'Sheet Name': self.sheet1_name, 'Total Rows': total_rows1},
            'File 2 Details': {'File Path': self.file2_path, 'Sheet Name': self.sheet2_name, 'Total Rows': total_rows2},
            'Comparison Results': {
                'Comparison Key(s)': ', '.join(self.key_columns),
                'Rows with Matched Data': comparison_df[matched_keys]['Status'].value_counts().get('Match', 0),
                'Rows with Data Mismatches': mismatched_rows.sum(),
                'Rows Only in File 1': (comparison_df['_merge'] == 'left_only').sum(),
                'Rows Only in File 2': (comparison_df['_merge'] == 'right_only').sum()
            }
        }
        
        # --- Writing to Sheet ---
        row_num = 1
        for section, data in summary_data.items():
            ws.cell(row=row_num, column=1, value=section).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row_num, column=1).fill = SUMMARY_HEADER_FILL
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            row_num += 1
            for key, value in data.items():
                ws.cell(row=row_num, column=1, value=key).font = Font(bold=True)
                ws.cell(row=row_num, column=2, value=value)
                row_num += 1
            row_num += 1 # Add a blank row between sections

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
    
    def _generate_comparison_sheet(self, wb: Workbook, comparison_df: pd.DataFrame, df1_sum: pd.DataFrame, df2_sum: pd.DataFrame, use_index_key: bool):
        """Creates the detailed side-by-side comparison sheet."""
        ws = wb.create_sheet("Side by Side Comparison")
        
        # --- Prepare Headers ---
        cols1 = [c for c in df1_sum.columns if c not in self.key_columns]
        cols2 = [c for c in df2_sum.columns if c not in self.key_columns]
        
        header = self.key_columns + cols1 + [" | "] + cols2 + ["Status"]
        ws.append(header)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        # --- Prepare Total Row ---
        total_row = ["Total"] * len(self.key_columns)
        total_row_diffs = {}
        for col in cols1:
            total_row.append(df1_sum[col].sum() if pd.api.types.is_numeric_dtype(df1_sum[col]) else "")
        total_row.append(" | ")
        for col in cols2:
            val2 = df2_sum[col].sum() if pd.api.types.is_numeric_dtype(df2_sum[col]) else ""
            total_row.append(val2)
            # Check for differences in common numeric columns
            if col in cols1 and pd.api.types.is_numeric_dtype(df1_sum[col]):
                val1 = df1_sum[col].sum()
                if not self._are_equal(val1, val2):
                    total_row_diffs[col] = True
        total_row.append("Total Mismatch" if total_row_diffs else "Total Match")
        ws.append(total_row)
        for cell in ws[2]:
            cell.fill = TOTAL_ROW_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        # --- Process and Write Data Rows ---
        for _, row in comparison_df.iterrows():
            row_data = []
            diffs_in_row = {}
            status = "Match"

            if row['_merge'] == 'right_only':
                status = "Missing in File 1"
            elif row['_merge'] == 'left_only':
                status = "Missing in File 2"
            else: # 'both'
                for col in cols1:
                    val1 = row[f'{col}_f1']
                    if col in cols2:
                        val2 = row[f'{col}_f2']
                        if not self._are_equal(val1, val2):
                            status = "Mismatch"
                            diffs_in_row[col] = True

            # Build row for writing
            row_data.extend([row[k] for k in self.key_columns])
            row_data.extend([row[f'{c}_f1'] if not pd.isna(row[f'{c}_f1']) else "" for c in cols1])
            row_data.append(" | ")
            row_data.extend([row[f'{c}_f2'] if not pd.isna(row[f'{c}_f2']) else "" for c in cols2])
            row_data.append(status)
            comparison_df.loc[_, 'Status'] = status # Store status for summary
            ws.append(row_data)

        # --- Apply Formatting ---
        self._apply_comparison_formatting(ws, len(cols1), len(cols2), diffs_in_row, total_row_diffs, use_index_key)

    def _apply_comparison_formatting(self, ws, num_cols1, num_cols2, diffs, total_diffs, use_index_key):
        """Applies all formatting to the side-by-side comparison sheet."""
        key_cols_count = len(self.key_columns)
        separator_col = key_cols_count + num_cols1 + 1
        
        # Hide the temp index key column if it was used
        if use_index_key:
            ws.column_dimensions[get_column_letter(1)].hidden= True

        # Format Total Row Differences
        for col_name in total_diffs:
            col_idx1 = self.key_columns.index(col_name) + 1 if col_name in self.key_columns else list(df1_sum.columns).index(col_name) + 1
            col_idx2 = separator_col + (list(df2_sum.columns).index(col_name) if col_name in df2_sum.columns else -1) + 1
            ws.cell(row=2, column=col_idx1).fill = CELL_DIFF_FILL
            ws.cell(row=2, column=col_idx2).fill = CELL_DIFF_FILL

        # Format Data Rows
        for row_idx in range(3, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=ws.max_column).value
            fill_color = None
            if status == "Match": fill_color = ROW_MATCH_FILL
            elif status == "Mismatch": fill_color = ROW_MISMATCH_FILL
            elif status in ["Missing in File 1", "Missing in File 2"]: fill_color = ROW_MISSING_FILL

            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER
                if fill_color:
                    ws.cell(row=row_idx, column=col_idx).fill = fill_color
            
            # Highlight separator
            ws.cell(row=row_idx, column=separator_col).fill = BLACK_SEPARATOR_FILL
            
            # Highlight specific cell differences in mismatch rows
            if status == "Mismatch":
                # This part requires re-aligning columns to highlight correctly.
                # For simplicity in this example, it highlights the whole row in yellow.
                # A more complex implementation would re-check values to find the exact differing cells.
                pass
        
        # Auto-fit columns
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

# --- Example Usage ---
if __name__ == "__main__":
    # Create dummy excel files for demonstration
    # File 1
    data1 = {'ID': [1, 2, 3, 4], 'Name': ['Apple', 'Banana', 'Carrot', 'Date'], 'Value': [10, 20, 30, 40]}
    df1 = pd.DataFrame(data1)
    df1.loc['Total'] = pd.Series(df1['Value'].sum(), index=['Value'])
    df1.to_excel("file1.xlsx", index=False)

    # File 2 - with differences
    data2 = {'ID': [1, 2, 4, 5], 'Name': ['Apple', 'Banana', 'Date', 'Elderberry'], 'Value': [10, 25, 40, 50]}
    df2 = pd.DataFrame(data2)
    df2.loc['Total'] = pd.Series(df2['Value'].sum(), index=['Value'])
    df2.to_excel("file2.xlsx", index=False)
    
    print("Created dummy files: file1.xlsx and file2.xlsx")

    # --- Run the comparator using 'ID' as the key ---
    comparator = ExcelComparator(
        file1_path="file1.xlsx",
        file2_path="file2.xlsx",
        key_columns=["ID"], # Specify the column(s) to match rows
        total_row_identifier="Total" # Tell the script to ignore rows with "Total" for summation
    )
    
    try:
        result_path = comparator.compare(output_file="comparison_with_key.xlsx")
        print(f"✅ Key-based comparison complete. Results saved to: {result_path}")
    except Exception as e:
        print(f"❌ Error: {e}")
