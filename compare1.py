import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue

def compare_excel_files():
    root = tk.Tk()
    root.withdraw()
    
    # File selection
    file1 = filedialog.askopenfilename(title="Select First Excel File")
    file2 = filedialog.askopenfilename(title="Select Second Excel File")
    if not file1 or not file2:
        return
    
    # Sheet type selection
    sheet_type = simpledialog.askstring(
        "Sheet Type", 
        "Enter sheet type:\n1. Same structure\n2. Different structures",
        initialvalue="1"
    )
    same_structure = sheet_type == "1" if sheet_type else True
    
    # Process files
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    
    # Sheet selection
    sheet1_name = simpledialog.askstring("Sheet Selection", "Enter sheet name for first file:", 
                                         initialvalue=wb1.sheetnames[0])
    sheet2_name = simpledialog.askstring("Sheet Selection", "Enter sheet name for second file:", 
                                         initialvalue=wb2.sheetnames[0])
    
    # Read data
    df1 = pd.read_excel(file1, sheet_name=sheet1_name)
    df2 = pd.read_excel(file2, sheet_name=sheet2_name)
    
    # Create comparison workbook
    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    
    # 1. Compare headers
    compare_headers(df1, df2, output_wb, same_structure)
    
    # 2. Compare row data
    compare_row_data(df1, df2, output_wb, same_structure)
    
    # 3. Numerical differences
    compare_numeric_values(df1, df2, output_wb)
    
    # Save results
    output_file = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if output_file:
        output_wb.save(output_file)
        messagebox.showinfo("Success", f"Comparison saved to:\n{output_file}")

def compare_headers(df1, df2, output_wb, same_structure=True):
    """Compare and highlight header differences"""
    headers1 = set(df1.columns)
    headers2 = set(df2.columns)
    common = headers1 & headers2
    unique1 = headers1 - headers2
    unique2 = headers2 - headers1
    
    # Create header comparison sheet
    ws = output_wb.create_sheet("Header Comparison")
    ws.append(["Header", "Status", "File 1 Presence", "File 2 Presence"])
    
    # Add common headers
    for header in sorted(common):
        ws.append([header, "Common", "✓", "✓"])
    
    # Add unique headers
    for header in sorted(unique1):
        ws.append([header, "Unique to File 1", "✓", ""])
        ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
    
    for header in sorted(unique2):
        ws.append([header, "Unique to File 2", "", "✓"])
        ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
    
    # Apply formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)

def compare_row_data(df1, df2, output_wb, same_structure=True):
    """Compare row-wise data and highlight differences"""
    # Align data
    df1, df2 = align_data(df1, df2, same_structure)
    
    # Create comparison sheets
    ws1 = output_wb.create_sheet("File1 Data")
    ws2 = output_wb.create_sheet("File2 Data")
    
    # Write headers
    for col_idx, header in enumerate(df1.columns, 1):
        ws1.cell(1, col_idx, header)
        ws2.cell(1, col_idx, header)
        ws1.cell(1, col_idx).font = Font(bold=True)
        ws2.cell(1, col_idx).font = Font(bold=True)
    
    # Compare and highlight cells
    for row_idx in range(len(df1)):
        for col_idx, col in enumerate(df1.columns, 1):
            val1 = df1.iloc[row_idx, col_idx-1]
            val2 = df2.iloc[row_idx, col_idx-1]
            
            # Write values
            ws1.cell(row_idx+2, col_idx, val1)
            ws2.cell(row_idx+2, col_idx, val2)
            
            # Highlight differences
            if not same_structure or (val1 != val2 and not (pd.isna(val1) and pd.isna(val2))):
                ws1.cell(row_idx+2, col_idx).fill = CELL_DIFF_FILL
                ws2.cell(row_idx+2, col_idx).fill = CELL_DIFF_FILL

def compare_numeric_values(df1, df2, output_wb):
    """Create numerical comparison table"""
    # Identify numeric columns
    num_cols = [col for col in df1.columns 
                if pd.api.types.is_numeric_dtype(df1[col]) and 
                pd.api.types.is_numeric_dtype(df2[col])]
    
    if not num_cols:
        return
    
    # Create sheet
    ws = output_wb.create_sheet("Numeric Comparison")
    headers = ["Column", "Row", "File1 Value", "File2 Value", "Absolute Diff", "Relative Diff"]
    ws.append(headers)
    
    # Apply header formatting
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Compare values
    for col in num_cols:
        for i in range(min(len(df1), len(df2))):
            val1 = df1[col].iloc[i]
            val2 = df2[col].iloc[i]
            
            if pd.isna(val1) or pd.isna(val2) or val1 == val2:
                continue
                
            abs_diff = abs(val1 - val2)
            rel_diff = abs_diff / max(abs(val1), abs(val2)) if max(abs(val1), abs(val2)) != 0 else float('inf')
            
            ws.append([col, i+1, val1, val2, abs_diff, rel_diff])
            
            # Highlight significant differences
            if rel_diff > 0.1:  > 10% difference
                for col_idx in range(1, 7):
                    ws.cell(ws.max_row, col_idx).fill = NUM_DIFF_FILL

def align_data(df1, df2, same_structure):
    """Align dataframes based on structure type"""
    if same_structure:
        # Pad with NaN to match lengths
        max_len = max(len(df1), len(df2))
        df1 = df1.reindex(range(max_len))
        df2 = df2.reindex(range(max_len))
        return df1, df2
    
    # Different structures - align by columns
    common_cols = list(set(df1.columns) & set(df2.columns))
    return df1[common_cols], df2[common_cols]

if __name__ == "__main__":
    compare_excel_files()
