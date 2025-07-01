import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from open极yxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue
ROW_MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light Green
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")       # Light Gray
TOTAL_FILL = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")        # Lavender

# Border style
THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

class ExcelComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Comparison Tool")
        self.root.geometry("900x700")
        self.root.configure(bg="#f0f2f5")
        
        # Initialize variables
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.sheet1_name = tk.StringVar()
        self.sheet2_name = tk.StringVar()
        self.status = tk.StringVar(value="Ready to compare files")
        self.df1 = None
        self.df2 = None
        
        # Create UI
        self.create_widgets()
        
    def create_widgets(self):
        # Header frame
        header_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        header_frame.pack(fill="x", side="top")
        
        header_label = tk.Label(
            header_frame, 
            text="Excel Data Comparison Tool", 
            font=("Arial", 20, "bold"), 
            fg="white", 
            bg="#2c3e50"
        )
        header_label.pack(pady=20)
        
        # Main content frame
        main_frame = tk.Frame(self.root, bg="#f0f2f5")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # File selection section
        file_frame = tk.LabelFrame(
            main_frame, 
            text="File Selection", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        file_frame.pack(fill="x", pady=(0, 15))
        
        # File 1
        file1_frame = tk.Frame(file_frame, bg="#f0f2f5")
        file1_frame.pack(fill="x", p极y=5)
        
        tk.Label(
            file1_frame, 
            text="File 1:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=10, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            file1_frame, 
            textvariable=self.file1_path, 
            width=50, 
            state="readonly",
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        tk.Button(
            file1_frame, 
            text="Browse", 
            command=lambda: self.select_file(1), 
            bg="#3498db", 
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left")
        
        # File 2
        file2_frame = tk.Frame(file_frame, bg="#f0f2f5")
        file2_frame.pack(fill="x", pady=5)
        
        tk.Label(
            file2_frame, 
            text="File 2:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=10, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            file2_frame, 
            textvariable=self.file2_path, 
            width=50, 
            state="readonly",
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        tk.Button(
            file2_frame, 
            text="Browse", 
            command=lambda: self.select_file(2), 
            bg="#3498db", 
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left")
        
        # Sheet selection section
        sheet_frame = tk.LabelFrame(
            main_frame, 
            text="Sheet Selection", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        sheet_frame.pack(fill="x", pady=(0, 15))
        
        # Sheet 1
        sheet1_frame = tk.Frame(sheet_frame, bg="#f0f2f5")
        sheet1_frame.pack(fill="x", pady=5)
        
        tk.Label(
            sheet1_frame, 
            text="File 1 Sheet:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=15, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            sheet1_frame, 
            textvariable=self.sheet1_name, 
            width=30, 
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        # Sheet 2
        sheet2_frame = tk.Frame(sheet_frame, bg="#f0f2f5")
        sheet2_frame.pack(fill="x", pady=5)
        
        tk.Label(
            sheet2_frame, 
            text="File 2 Sheet:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=15, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            sheet2_frame, 
            textvariable=self.sheet2_name, 
            width=30, 
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        # Options section
        options_frame = tk.LabelFrame(
            main_frame, 
            text="Comparison Options", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        options_frame.pack(fill="x", pady=(0, 15))
        
        self.highlight_missing = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight missing columns", 
            variable=self.highlight_missing, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.highlight_cell_diffs = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight cell differences", 
            variable=self.highlight_cell_diffs, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.highlight_row_matches = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight row matches/mismatches", 
            variable=self.highlight_row_matches, 
            bg="#f0f2f5", 
            font=("Arial", 极10)
        ).pack(anchor="w", pady=3)
        
        self.create_num_table = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Create numerical differences table", 
            variable=self.create_num_table, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        # Action buttons
        button_frame = tk.Frame(main_frame, bg="#f0f2f5")
        button_frame.pack(fill="x", pady=20)
        
        compare_btn = tk.Button(
            button_frame, 
            text="Compare Excel Files", 
            command=self.compare_files, 
            bg="#27ae60", 
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=20
        )
        compare_btn.pack(pady=10)
        
        # Status bar
        status_frame = tk.Frame(self.root, bg="#e0e0e0", height=30)
        status_frame.pack(fill="x", side="bottom")
        
        tk.Label(
            status_frame, 
            textvariable=self.status, 
            font=("Arial", 10), 
            bg="#e0e0e0", 
            anchor="w"
        ).pack(side="left", padx=10)
        
        # Legend
        legend_frame = tk.Frame(main_frame, bg="#f0f2f5")
        legend_frame.pack(fill="x", pady=10)
        
        tk.Label(
            legend_frame, 
            text="Highlight Legend:", 
            font=("Arial", 10, "bold"), 
            bg="#f0f2f5"
        ).pack(anchor="w")
        
        legend_inner = tk.Frame(legend_frame, bg="#f0f2f5")
        legend_inner.pack(fill="x", pady=5)
        
        # Header diff legend
        header_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        header_legend.pack(side="left", padx=10)
        tk.Label(
            header_legend, 
            text="    ", 
            bg="#FFD700", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            header_legend, 
            text="Column Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Cell diff legend
        cell_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        cell_legend.pack(side="left", padx=10)
        tk.Label(
            cell_legend, 
            text="    ", 
            bg="#FF6347", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            cell_legend, 
            text="Value Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Row match legend
        row_match_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        row_match_legend.pack(side="left", padx=10)
        tk.Label(
            row_match_legend, 
            text="    ", 
            bg="#90EE90", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            row_match_legend, 
            text="Matched Rows", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Row missing legend
        row_missing_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        row_missing_legend.pack(side="left", padx=10)
        tk.Label(
            row_missing_legend, 
            text="    ", 
            bg="#D3D3D3", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            row_missing_legend, 
            text="Unmatched Rows", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Total legend
        total_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        total_legend.pack(side="left", padx=10)
        tk.Label(
            total_legend, 
            text="    ", 
            bg="#E6E6FA", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            total_legend, 
            text="Total Row", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
    
    def select_file(self, file_num):
        file_path = filedialog.askopenfilename(
            title=f"Select Excel File {file_num}",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            if file_num == 1:
                self.file1_path.set(file_path)
                try:
                    wb = load_workbook(file_path, read_only=True)
                    self.sheet1_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet1_name.set("Sheet1")
            else:
                self.file2_path.set(file_path)
                try:
                    wb = load_workbook(file_path, read_only=True)
                    self.sheet2_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet2_name.set("Sheet1")
    
    def are_equal(self, a, b):
        """Check if two values are equal, handling NaN cases"""
        if pd.isna(a) and pd.isna(b):
            return True
        if pd.isna(a) or pd.isna(b):
            return False
        return a == b
    
    def is_date(self, value):
        """Check if a value is a date"""
        return isinstance(value, (datetime, pd.Timestamp))
    
    def get_string_columns(self, df):
        """Get all string columns that are not dates"""
        string_cols = []
        for col in df.columns:
            # Skip date columns
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                continue
            
            # Include string columns and object columns that are likely strings
            if pd.api.types.is_string_dtype(df[col]) or \
               (pd.api.types.is_object_dtype(df[col]) and 
                all((isinstance(x, str) or pd.isna(x)) for x in df[col].head(100))):
                string_cols.append(col)
        return string_cols
    
    def create_side_by_side_sheet(self, df1, df2, output_wb):
        """Create side-by-side comparison sheet with row matching column and totals"""
        # Create sheet
        ws = output_wb.create_sheet("Side by Side Comparison")
        
        # Get common columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        
        # Get all string columns (excluding dates)
        str_cols1 = self.get_string_columns(df1)
        str_cols2 = self.get_string_columns(df2)
        all_str_cols = list(set(str_cols1) | set(str_cols2))
        
        # Get numerical columns
        num_cols1 = [col for col in df1.columns if pd.api.types.is_numeric_dtype(df1[col])]
        num_cols2 = [col for col in df2.columns if pd.api.types.is_numeric_dtype(df2[col])]
        common_num_cols = list(set(num_cols1) & set(num_cols2))
        
        # Create concatenation keys using all string columns
        concat_keys1 = {}
        concat_keys2 = {}
        
        # Create keys for df1
        for idx, row in df1.iterrows():
            key_parts = []
            for col in all_str_cols:
                if col in df1.columns:
                    val = row[col]
                    if pd.notna(val) and not self.is_date(val):
                        key_parts.append(str(val))
            concat_keys1[idx] = "_".join(key_parts) if key_parts else None
        
        # Create keys for df2
        for idx, row in df2.iterrows():
            key_parts = []
            for col in all_str_cols:
                if col in df2.columns:
                    val = row[col]
                    if pd.notna(val) and not self.is_date(val):
                        key_parts.append(str(val))
            concat_keys2[idx] = "_".join(key_parts) if key_parts else None
        
        # Create sets of keys for matching
        keys1_set = set(concat_keys1.values())
        keys2_set = set(concat_keys2.values())
        
        # Create match status for each row in File1
        df1_match_status = []
        for idx in df1.index:
            key = concat_keys1[idx]
            match_found = False
            
            # Check for string key match
            if key and key in keys2_set:
                match_found = True
            # Check for numerical match if no string match
            elif common_num_cols:
                for idx2 in df2.index:
                    num_match = True
                    for col in common_num_cols:
                        if not self.are_equal(df1.at[idx, col], df2.at[idx2, col]):
                            num_match = False
                            break
                    if num_match:
                        match_found = True
                        break
                        
            df1_match_status.append("Matched" if match_found else "Not Matched")
        
        # Add Match Status column to File1
        df1['Match Status'] = df1_match_status
        
        # Write headers
        header_row = list(df1.columns) + list(df2.columns)
        ws.append(header_row)
        
        # Apply header styling
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        
        # Add totals row
        totals_row = []
        
        # File1 totals
        for col in df1.columns:
            if col in num_cols1:
                totals_row.append(df1[col].sum())
            else:
                totals_row.append("")
        
        # File2 totals
        for col in df2.columns:
            if col in num_cols2:
                totals_row.append(df2[col].sum())
            else:
                totals_row.append("")
        
        ws.append(totals_row)
        
        # Apply styling to totals row
        for col_idx in range(1, len(totals_row) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        
        # Write data row by row
        max_rows = max(len(df1), len(df2))
        for i in range(max_rows):
            row_data = []
            
            # Add File1 data if exists
            if i < len(df1):
                row_data.extend(df1.iloc[i].tolist())
            else:
                row_data.extend([""] * len(df1.columns))
            
            # Add File2 data if exists
            if i < len(df2):
                row_data.extend(df2.iloc[i].tolist())
            else:
                row_data.extend([""] * len(df2.columns))
            
            ws.append(row_data)
            
            # Apply row matching highlighting
            if self.highlight_row_matches.get():
                if i < len(df1):
                    match_status = df1.iloc[i]['Match Status']
                    
                    # Calculate column positions
                    file1_start_col = 1
                    file1_end_col = len(df1.columns)
                    
                    # Apply styling
                    if match_status == "Matched":
                        for col_idx in range(file1_start_col, file1_end_col + 1):
                            ws.cell(row=i+3, column=col_idx).fill = ROW_MATCH_FILL
                    else:
                        for col_idx in range(file1_start_col, file1_end_col + 1):
                            ws.cell(row=i+3, column=col_idx).fill = ROW_MISSING_FILL
            
            # Apply cell difference highlighting
            if self.highlight_cell_diffs.get() and i < len(df1) and i < len(df2):
                for col in common_cols:
                    val1 = df1.at[i, col]
                    val2 = df2.at[i, col]
                    
                    if not self.are_equal(val1, val2):
                        col_idx1 = list(df1.columns).index(col) + 1
                        col_idx2 = list(df2.columns).index(col) + len(df1.columns) + 1
                        
                        ws.cell(row=i+3, column=col_idx1).fill = CELL_DIFF_FILL
                        ws.cell(row=i+3, column=col_idx2).fill = CELL_DIFF_FILL
            
            # Apply borders
            for col_idx in range(1, len(header_row) + 1):
                ws.cell(row=i+3, column=col_idx).border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, len(header_row) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = "A3"
        
        matched_count = len(df1[df1['Match Status'] == "Matched"])
        unmatched_count = len(df1[df1['Match Status'] == "Not Matched"])
        return matched_count, unmatched_count
    
    def compare_files(self):
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()
        sheet1 = self.sheet1_name.get()
        sheet2 = self.sheet2_name.get()
        
        if not file1 or not file2:
            messagebox.showerror("Error", "Please select both Excel files")
            return
        
        self.status.set("Reading files...")
        self.root.update()
        
        try:
            # Read data
            df1 = pd.read_excel(file1, sheet_name=sheet1)
            df2 = pd.read_excel(file2, sheet_name=sheet2)
            
            # Create comparison workbook
            output_wb = Workbook()
            output_wb.remove(output_wb.active)
            
            # 1. Compare headers
            self.compare_headers(df1, df2, output_wb)
            
            # 2. Create side-by-side comparison sheet
            matched_count, unmatched_count = self.create_side_by_side_sheet(df1, df2, output_wb)
            
            # 3. Row matching analysis
            self.analyze_row_matches(df1, df2, output_wb, matched_count, unmatched_count)
            
            # 4. Numerical differences
            if self.create_num_table.get():
                self.compare_numeric_values(df1, df2, output_wb)
            
            # Save results
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Comparison Results"
            )
            
            if output_file:
                output_wb.save(output_file)
                self.status.set(f"Comparison saved to: {output_file}")
                messagebox.showinfo("Success", f"Comparison saved successfully!\n{output_file}")
            else:
                self.status.set("Comparison canceled")
                
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.status.set("Error occurred - see details in message")
    
    def compare_headers(self, df1, df2, output_wb):
        """Compare and highlight header differences"""
        headers1 = set(df1.columns)
        headers2 = set(df2.columns)
        common = headers1 & headers2
        unique1 = headers1 - headers2
        unique2 = headers2 - headers1
        
        # Create header comparison sheet
        ws = output_wb.create_sheet("Header Comparison")
        ws.append(["Header", "Status", "File 1 Presence", "File 2 Presence"])
        
        # Add common headers
        for header in sorted(common):
            ws.append([header, "Common", "✓", "✓"])
        
        # Add unique headers if option is enabled
        if self.highlight_missing.get():
            for header in sorted(unique1):
                ws.append([header, "Unique to File 1", "✓", ""])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
            
            for header in sorted(unique2):
                ws.append([header, "Unique to File 2", "", "✓"])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
        
        # Apply formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(bold=(cell.row == 1))
                cell.border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, 5):  # We have 4 columns in this sheet
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def analyze_row_matches(self, df1, df2, output_wb, matched_count, unmatched_count):
        """Analyze and highlight row matches between files"""
        # Create row matching sheet
        ws = output_wb.create_sheet("Row Matching Analysis")
        
        # Summary section
        ws.append(["Row Matching Summary"])
        ws.append(["", ""])
        ws.append(["Total Rows in File1", len(df1)])
        ws.append(["Total Rows in File2", len(df2)])
        ws.append(["Matched Rows in File1", matched_count])
        ws.append(["Unmatched Rows in File1", unmatched_count])
        ws.append([""])
        
        # Key generation method
        ws.append(["Matching Logic:"])
        ws.append(["1. Match based on concatenated string columns (non-date)"])
        ws.append(["2. If no string match, match based on numerical columns"])
        ws.append(["3. Rows considered matched if any row in File1 matches any row in File2"])
        
        # Apply styling to summary
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)
        
        for row in ws.iter_rows(min_row=3, max_row=6):
            for cell in row:
                cell.font = Font(bold=(cell.column == 1))
        
        # Auto-size columns
        for col_idx in range(1, 3):  # We have 2 columns in this sheet
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def compare_numeric_values(self, df1, df2, output_wb):
        """Create numerical comparison table for common numeric columns"""
        # Identify common numeric columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        num_cols = [col for col in common_cols 
                    if pd.api.types.is_numeric_dtype(df1[col]) and 
                    pd.api.types.is_numeric_dtype(df2[col])]
        
        if not num_cols:
            return
        
        # Create sheet
        ws = output_wb.create_sheet("Numeric Comparison")
        headers = ["Column", "Row", "File1 Value", "File2 Value", "Absolute Diff", "Relative Diff"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        # Compare values
        for col in num_cols:
            for i in range(min(len(df1), len(df2))):
                val1 = df1[col].iloc[i]
                val2 = df2[col].iloc[i]
                
                if pd.isna(val1) or pd.isna(val2) or val1 == val2:
                    continue
                    
                abs_diff = abs(val1 - val2)
                rel_diff = abs_diff / max(abs(val1), abs(val2)) if max(abs(val1), abs(val2)) != 0 else float('inf')
                
                ws.append([col, i+1, val1, val2, abs_diff, rel_diff])
                
                # Highlight significant differences (>10%)
                if rel_diff > 0.1:
                    for col_idx in range(1, 7):
                        cell = ws.cell(ws.max_row, col_idx)
                        cell.fill = NUM_DIFF_FILL
                
                # Apply borders
                for col_idx in range(1, 7):
                    cell = ws.cell(ws.max_row, col_idx)
                    cell.border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, 7):  # We have 6 columns in this sheet
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparator(root)
    root.mainloop()
