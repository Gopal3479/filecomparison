import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue
ROW_MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light Green
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")       # Light Gray

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

class ExcelComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Advanced Comparison Tool")
        self.root.geometry("900x700")
        self.status = tk.StringVar(value="Ready")

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.sheet1_name = tk.StringVar()
        self.sheet2_name = tk.StringVar()
        
        self.build_ui()

    def build_ui(self):
        tk.Label(self.root, text="Excel Comparison Tool", font=("Arial", 18, "bold")).pack(pady=10)
        frame = tk.Frame(self.root)
        frame.pack(pady=10)

        tk.Label(frame, text="File 1:").grid(row=0, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.file1_path, width=60).grid(row=0, column=1)
        tk.Button(frame, text="Browse", command=lambda: self.browse_file(1)).grid(row=0, column=2)

        tk.Label(frame, text="Sheet 1:").grid(row=1, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.sheet1_name, width=20).grid(row=1, column=1, sticky="w")

        tk.Label(frame, text="File 2:").grid(row=2, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.file2_path, width=60).grid(row=2, column=1)
        tk.Button(frame, text="Browse", command=lambda: self.browse_file(2)).grid(row=2, column=2)

        tk.Label(frame, text="Sheet 2:").grid(row=3, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.sheet2_name, width=20).grid(row=3, column=1, sticky="w")

        tk.Button(self.root, text="Compare Files", command=self.compare_files, bg="green", fg="white").pack(pady=20)

        tk.Label(self.root, textvariable=self.status, fg="blue").pack(pady=5)

    def browse_file(self, file_number):
        path = filedialog.askopenfilename(filetypes=[("Excel Files","*.xlsx *.xls")])
        if path:
            if file_number == 1:
                self.file1_path.set(path)
                try:
                    wb = load_workbook(path, read_only=True)
                    self.sheet1_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet1_name.set("Sheet1")
            else:
                self.file2_path.set(path)
                try:
                    wb = load_workbook(path, read_only=True)
                    self.sheet2_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet2_name.set("Sheet1")

    def compare_files(self):
        try:
            self.status.set("Loading files...")
            self.root.update()

            df1 = pd.read_excel(self.file1_path.get(), sheet_name=self.sheet1_name.get())
            df2 = pd.read_excel(self.file2_path.get(), sheet_name=self.sheet2_name.get())

            output_wb = Workbook()
            output_wb.remove(output_wb.active)

            self.compare_headers(df1, df2, output_wb)
            self.create_side_by_side_sheet(df1, df2, output_wb)
            self.compare_numeric_values(df1, df2, output_wb)

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                      filetypes=[("Excel Files","*.xlsx")])
            if save_path:
                output_wb.save(save_path)
                self.status.set(f"Saved: {save_path}")
                messagebox.showinfo("Done", f"Comparison saved at:\n{save_path}")
            else:
                self.status.set("Cancelled")

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.status.set("Error during comparison")

    def compare_headers(self, df1, df2, wb):
        ws = wb.create_sheet("Header Comparison")
        ws.append(["Header", "In File1", "In File2", "Status"])

        headers1 = set(df1.columns)
        headers2 = set(df2.columns)
        all_headers = sorted(headers1 | headers2)

        for header in all_headers:
            status = "Common" if header in headers1 and header in headers2 else (
                     "Missing in File2" if header in headers1 else "Missing in File1")
            ws.append([header,
                       "Yes" if header in headers1 else "No",
                       "Yes" if header in headers2 else "No",
                       status])
            if status != "Common":
                for cell in ws[ws.max_row]:
                    cell.fill = HEADER_DIFF_FILL

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    def create_side_by_side_sheet(self, df1, df2, wb):
        ws = wb.create_sheet("Side by Side Comparison")

        common_cols = sorted(set(df1.columns).intersection(set(df2.columns)))
        ws.append([f"File1_{c}" for c in common_cols] + [f"File2_{c}" for c in common_cols] + ["Match"])

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        max_len = max(len(df1), len(df2))
        for i in range(max_len):
            row = []
            match_status = "Matched"
            for col in common_cols:
                val1 = df1.iloc[i][col] if i < len(df1) else None
                val2 = df2.iloc[i][col] if i < len(df2) else None
                row.append(val1)
                row.append(val2)

                if pd.isna(val1) and pd.isna(val2):
                    continue
                if val1 != val2:
                    match_status = "Different"

            row.append(match_status)
            ws.append(row)

            row_idx = ws.max_row
            for j, col in enumerate(common_cols):
                val1 = df1.iloc[i][col] if i < len(df1) else None
                val2 = df2.iloc[i][col] if i < len(df2) else None
                if val1 != val2:
                    ws.cell(row=row_idx, column=2*j+1).fill = CELL_DIFF_FILL
                    ws.cell(row=row_idx, column=2*j+2).fill = CELL_DIFF_FILL

            match_cell = ws.cell(row=row_idx, column=2*len(common_cols)+1)
            match_cell.fill = ROW_MATCH_FILL if match_status=="Matched" else CELL_DIFF_FILL

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    def compare_numeric_values(self, df1, df2, wb):
        ws = wb.create_sheet("Numeric Differences")

        common_cols = [col for col in df1.columns if col in df2.columns 
                       and pd.api.types.is_numeric_dtype(df1[col]) 
                       and pd.api.types.is_numeric_dtype(df2[col])]
        
        ws.append(["Column", "Row", "File1", "File2", "Abs Diff", "Rel Diff"])

        for col in common_cols:
            for idx in range(min(len(df1), len(df2))):
                v1 = df1[col].iloc[idx]
                v2 = df2[col].iloc[idx]
                if pd.isna(v1) or pd.isna(v2) or v1==v2:
                    continue
                abs_diff = abs(v1 - v2)
                rel_diff = abs_diff / max(abs(v1), abs(v2)) if max(abs(v1),abs(v2)) else float("inf")
                ws.append([col, idx+1, v1, v2, abs_diff, rel_diff])

                if rel_diff > 0.1:
                    for c in ws[ws.max_row]:
                        c.fill = NUM_DIFF_FILL

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

if __name__ == "__main__":
    root = tk.Tk()
    ExcelComparator(root)
    root.mainloop()
