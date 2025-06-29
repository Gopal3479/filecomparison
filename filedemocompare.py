import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import re

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue
ROW_MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light Green
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
TOTAL_FILL = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")        # Lavender
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")       # Light Gray

# Border style
THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

class ExcelComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Excel Data Comparison Tool")
        self.root.geometry("900x700")
        self.root.configure(bg="#f0f2f5")
        
        # Initialize variables
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.sheet1_name = tk.StringVar()
        self.sheet2_name = tk.StringVar()
        self.status = tk.StringVar(value="Ready to compare files")
        self.df1 = None
        self.df2 = None
        self.side_by_side_df = None
        
        # Create UI
        self.create_widgets()
        
    def create_widgets(self):
        # Header frame
        header_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        header_frame.pack(fill="x", side="top")
        
        header_label = tk.Label(
            header_frame, 
            text="Advanced Excel Data Comparison Tool", 
            font=("Arial", 20, "bold"), 
            fg="white", 
            bg="#2c3e50"
        )
        header_label.pack(pady=20)
        
        # Main content frame
        main_frame = tk.Frame(self.root, bg="#f0f2f5")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # File selection section
        file_frame = tk.LabelFrame(
            main_frame, 
            text="File Selection", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        file_frame.pack(fill="x", pady=(0, 15))
        
        # File 1
        file1_frame = tk.Frame(file_frame, bg="#f0f2f5")
        file1_frame.pack(fill="x", pady=5)
        
        tk.Label(
            file1_frame, 
            text="File 1:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=10, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            file1_frame, 
            textvariable=self.file1_path, 
            width=50, 
            state="readonly",
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        tk.Button(
            file1_frame, 
            text="Browse", 
            command=lambda: self.select_file(1), 
            bg="#3498db", 
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left")
        
        # File 2
        file2_frame = tk.Frame(file_frame, bg="#f0f2f5")
        file2_frame.pack(fill="x", pady=5)
        
        tk.Label(
            file2_frame, 
            text="File 2:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=10, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            file2_frame, 
            textvariable=self.file2_path, 
            width=50, 
            state="readonly",
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        tk.Button(
            file2_frame, 
            text="Browse", 
            command=lambda: self.select_file(2), 
            bg="#3498db", 
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left")
        
        # Sheet selection section
        sheet_frame = tk.LabelFrame(
            main_frame, 
            text="Sheet Selection", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        sheet_frame.pack(fill="x", pady=(0, 15))
        
        # Sheet 1
        sheet1_frame = tk.Frame(sheet_frame, bg="#f0f2f5")
        sheet1_frame.pack(fill="x", pady=5)
        
        tk.Label(
            sheet1_frame, 
            text="File 1 Sheet:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=15, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            sheet1_frame, 
            textvariable=self.sheet1_name, 
            width=30, 
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        # Sheet 2
        sheet2_frame = tk.Frame(sheet_frame, bg="#f0f2f5")
        sheet2_frame.pack(fill="x", pady=5)
        
        tk.Label(
            sheet2_frame, 
            text="File 2 Sheet:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=15, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            sheet2_frame, 
            textvariable=self.sheet2_name, 
            width=30, 
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        # Options section
        options_frame = tk.LabelFrame(
            main_frame, 
            text="Comparison Options", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        options_frame.pack(fill="x", pady=(0, 15))
        
        self.highlight_missing = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight missing columns", 
            variable=self.highlight_missing, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.highlight_cell_diffs = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight cell differences", 
            variable=self.highlight_cell_diffs, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.highlight_row_matches = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight row matches/mismatches", 
            variable=self.highlight_row_matches, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.create_num_table = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Create numerical differences table", 
            variable=self.create_num_table, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.create_totals = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Show totals for numerical columns", 
            variable=self.create_totals, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        # Action buttons
        button_frame = tk.Frame(main_frame, bg="#f0f2f5")
        button_frame.pack(fill="x", pady=20)
        
        compare_btn = tk.Button(
            button_frame, 
            text="Compare Excel Files", 
            command=self.compare_files, 
            bg="#27ae60", 
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=20
        )
        compare_btn.pack(pady=10)
        
        # Status bar
        status_frame = tk.Frame(self.root, bg="#e0e0e0", height=30)
        status_frame.pack(fill="x", side="bottom")
        
        tk.Label(
            status_frame, 
            textvariable=self.status, 
            font=("Arial", 10), 
            bg="#e0e0e0", 
            anchor="w"
        ).pack(side="left", padx=10)
        
        # Legend
        legend_frame = tk.Frame(main_frame, bg="#f0f2f5")
        legend_frame.pack(fill="x", pady=10)
        
        tk.Label(
            legend_frame, 
            text="Highlight Legend:", 
            font=("Arial", 10, "bold"), 
            bg="#f0f2f5"
        ).pack(anchor="w")
        
        legend_inner = tk.Frame(legend_frame, bg="#f0f2f5")
        legend_inner.pack(fill="x", pady=5)
        
        # Header diff legend
        header_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        header_legend.pack(side="left", padx=10)
        tk.Label(
            header_legend, 
            text="    ", 
            bg="#FFD700", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            header_legend, 
            text="Column Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Cell diff legend
        cell_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        cell_legend.pack(side="left", padx=10)
        tk.Label(
            cell_legend, 
            text="    ", 
            bg="#FF6347", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            cell_legend, 
            text="Value Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Row match legend
        row_match_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        row_match_legend.pack(side="left", padx=10)
        tk.Label(
            row_match_legend, 
            text="    ", 
            bg="#90EE90", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            row_match_legend, 
            text="Matched Rows", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Row missing legend
        row_missing_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        row_missing_legend.pack(side="left", padx=10)
        tk.Label(
            row_missing_legend, 
            text="    ", 
            bg="#D3D3D3", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            row_missing_legend, 
            text="Unmatched Rows", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Total legend
        total_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        total_legend.pack(side="left", padx=10)
        tk.Label(
            total_legend, 
            text="    ", 
            bg="#E6E6FA", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            total_legend, 
            text="Total Row", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
    
    def select_file(self, file_num):
        file_path = filedialog.askopenfilename(
            title=f"Select Excel File {file_num}",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            if file_num == 1:
                self.file1_path.set(file_path)
                try:
                    wb = load_workbook(file_path, read_only=True)
                    self.sheet1_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet1_name.set("Sheet1")
            else:
                self.file2_path.set(file_path)
                try:
                    wb = load_workbook(file_path, read_only=True)
                    self.sheet2_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet2_name.set("Sheet1")
    
    def are_equal(self, a, b):
        """Check if two values are equal, handling NaN cases"""
        if pd.isna(a) and pd.isna(b):
            return True
        if pd.isna(a) or pd.isna(b):
            return False
        return a == b
    
    def is_date(self, value):
        """Check if a value is a date"""
        return isinstance(value, (datetime, pd.Timestamp))
    
    def get_string_columns(self, df):
        """Get all string columns that are not dates"""
        string_cols = []
        for col in df.columns:
            # Skip date columns
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                continue
            
            # Include string columns and object columns that are likely strings
            if pd.api.types.is_string_dtype(df[col]) or \
               (pd.api.types.is_object_dtype(df[col]) and 
                all(isinstance(x, str) or pd.isna(x) for x in df[col].head(100))):
                string_cols.append(col)
        return string_cols
    
    def create_side_by_side_sheet(self, df1, df2, output_wb):
        """Create side-by-side comparison sheet with totals and row matching"""
        # Create sheet
        ws = output_wb.create_sheet("Side by Side Comparison")
        
        # Get common columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        
        # Get all string columns (excluding dates)
        str_cols1 = self.get_string_columns(df1)
        str_cols2 = self.get_string_columns(df2)
        all_str_cols = list(set(str_cols1) | set(str_cols2))
        
        # Create concatenation keys using all string columns
        concat_keys1 = {}
        concat_keys2 = {}
        
        # Create keys for df1
        for idx, row in df1.iterrows():
            key_parts = []
            for col in all_str_cols:
                if col in df1.columns:
                    val = row[col]
                    if pd.notna(val) and not self.is_date(val):
                        key_parts.append(str(val))
            concat_keys1[idx] = "_".join(key_parts) if key_parts else None
        
        # Create keys for df2
        for idx, row in df2.iterrows():
            key_parts = []
            for col in all_str_cols:
                if col in df2.columns:
                    val = row[col]
                    if pd.notna(val) and not self.is_date(val):
                        key_parts.append(str(val))
            concat_keys2[idx] = "_".join(key_parts) if key_parts else None
        
        # Create sets of keys for matching
        keys1_set = set(concat_keys1.values())
        keys2_set = set(concat_keys2.values())
        
        # Prepare data for side-by-side comparison
        side_by_side_data = []
        
        # Create a mapping of keys to row indices
        key_to_df1 = {}
        for idx, key in concat_keys1.items():
            if key is not None and key != "":
                key_to_df1.setdefault(key, []).append(idx)
        
        key_to_df2 = {}
        for idx, key in concat_keys2.items():
            if key is not None and key != "":
                key_to_df2.setdefault(key, []).append(idx)
        
        # Find all unique keys
        all_keys = set(list(concat_keys1.values()) + list(concat_keys2.values()))
        all_keys.discard(None)
        all_keys.discard("")
        
        # Process each key to find matches
        matched_rows = []
        for key in all_keys:
            df1_rows = key_to_df1.get(key, [])
            df2_rows = key_to_df2.get(key, [])
            
            # If we have at least one row in both files, it's a match
            if df1_rows and df2_rows:
                # Take the first matching row from each file
                matched_rows.append((df1_rows[0], df2_rows[0]))
        
        # Create sets of matched indices
        matched_df1_indices = set()
        matched_df2_indices = set()
        for df1_idx, df2_idx in matched_rows:
            matched_df1_indices.add(df1_idx)
            matched_df2_indices.add(df2_idx)
        
        # Create list of unmatched rows
        unmatched_df1 = [idx for idx in df1.index if idx not in matched_df1_indices]
        unmatched_df2 = [idx for idx in df2.index if idx not in matched_df2_indices]
        
        # Prepare side-by-side data
        # First add matched rows
        for df1_idx, df2_idx in matched_rows:
            row_data = []
            # Add File1 data
            for col in df1.columns:
                row_data.append(df1.at[df1_idx, col])
            # Add File2 data
            for col in df2.columns:
                row_data.append(df2.at[df2_idx, col])
            row_data.append("Matched")
            side_by_side_data.append(row_data)
        
        # Then add unmatched from File1
        for idx in unmatched_df1:
            row_data = []
            # Add File1 data
            for col in df1.columns:
                row_data.append(df1.at[idx, col])
            # Add blank for File2
            for _ in df2.columns:
                row_data.append(None)
            row_data.append("Not Matched (File1)")
            side_by_side_data.append(row_data)
        
        # Then add unmatched from File2
        for idx in unmatched_df2:
            row_data = []
            # Add blank for File1
            for _ in df1.columns:
                row_data.append(None)
            # Add File2 data
            for col in df2.columns:
                row_data.append(df2.at[idx, col])
            row_data.append("Not Matched (File2)")
            side_by_side_data.append(row_data)
        
        # Create DataFrame for side-by-side view
        columns = [f"File1_{col}" for col in df1.columns] + \
                  [f"File2_{col}" for col in df2.columns] + \
                  ["Match Status"]
        self.side_by_side_df = pd.DataFrame(side_by_side_data, columns=columns)
        
        # Write headers
        header_row = ["File1"] * len(df1.columns) + ["File2"] * len(df2.columns) + [""]
        ws.append(header_row)
        
        col_names = [col for col in df1.columns] + [col for col in df2.columns] + ["Match Status"]
        ws.append(col_names)
        
        # Apply header styling
        for row in ws.iter_rows(min_row=1, max_row=2, max_col=len(col_names)):
            for cell in row:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
        
        # Merge header cells
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df1.columns))
        ws.merge_cells(start_row=1, start_column=len(df1.columns)+1, 
                      end_row=1, end_column=len(df1.columns)+len(df2.columns))
        
        # Set alignment for merged headers
        ws.cell(row=1, column=1).value = "File1"
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=len(df1.columns)+1).value = "File2"
        ws.cell(row=1, column=len(df1.columns)+1).alignment = Alignment(horizontal='center')
        
        # Add totals row if enabled
        if self.create_totals.get():
            totals_row = []
            
            # File1 totals
            for col in df1.columns:
                if pd.api.types.is_numeric_dtype(df1[col]):
                    totals_row.append(df1[col].sum())
                else:
                    totals_row.append("")
            
            # File2 totals
            for col in df2.columns:
                if pd.api.types.is_numeric_dtype(df2[col]):
                    totals_row.append(df2[col].sum())
                else:
                    totals_row.append("")
            
            totals_row.append("Totals")
            ws.append(totals_row)
            
            # Apply styling to totals row
            totals_row_num = ws.max_row
            for col_idx in range(1, len(totals_row) + 1):
                cell = ws.cell(row=totals_row_num, column=col_idx)
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)
                cell.border = THIN_BORDER
        
        # Write data
        for _, row in self.side_by_side_df.iterrows():
            ws.append(row.tolist())
        
        # Apply styling and formatting
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), 3):
            # Skip totals row
            if self.create_totals.get() and row_idx == 3:
                continue
                
            match_status = ws.cell(row=row_idx, column=len(col_names)).value
            
            # Apply row matching highlighting
            if self.highlight_row_matches.get():
                if match_status == "Matched":
                    for cell in row:
                        cell.fill = ROW_MATCH_FILL
                elif "Not Matched" in match_status:
                    for cell in row:
                        cell.fill = ROW_MISSING_FILL
            
            # Apply cell difference highlighting
            if self.highlight_cell_diffs.get() and match_status == "Matched":
                # Only compare common columns for matched rows
                for col_idx, col_name in enumerate(df1.columns, 1):
                    if col_name in common_cols:
                        file1_val = ws.cell(row=row_idx, column=col_idx).value
                        file2_val = ws.cell(row=row_idx, column=col_idx + len(df1.columns)).value
                        
                        if not self.are_equal(file1_val, file2_val):
                            ws.cell(row=row_idx, column=col_idx).fill = CELL_DIFF_FILL
                            ws.cell(row=row_idx, column=col_idx + len(df1.columns)).fill = CELL_DIFF_FILL
            
            # Apply borders
            for cell in row:
                cell.border = THIN_BORDER
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = "C3"
        
        return len(matched_rows), len(unmatched_df1), len(unmatched_df2)
    
    def compare_files(self):
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()
        sheet1 = self.sheet1_name.get()
        sheet2 = self.sheet2_name.get()
        
        if not file1 or not file2:
            messagebox.showerror("Error", "Please select both Excel files")
            return
        
        self.status.set("Reading files...")
        self.root.update()
        
        try:
            # Read data
            df1 = pd.read_excel(file1, sheet_name=sheet1)
            df2 = pd.read_excel(file2, sheet_name=sheet2)
            
            # Create comparison workbook
            output_wb = Workbook()
            output_wb.remove(output_wb.active)
            
            # 1. Compare headers
            self.compare_headers(df1, df2, output_wb)
            
            # 2. Create side-by-side comparison sheet
            matched_count, unmatched1_count, unmatched2_count = self.create_side_by_side_sheet(df1, df2, output_wb)
            
            # 3. Row matching analysis
            self.analyze_row_matches(df1, df2, output_wb, matched_count, unmatched1_count, unmatched2_count)
            
            # 4. Numerical differences
            if self.create_num_table.get():
                self.compare_numeric_values(df1, df2, output_wb)
            
            # Save results
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Comparison Results"
            )
            
            if output_file:
                output_wb.save(output_file)
                self.status.set(f"Comparison saved to: {output_file}")
                messagebox.showinfo("Success", f"Comparison saved successfully!\n{output_file}")
            else:
                self.status.set("Comparison canceled")
                
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.status.set("Error occurred - see details in message")
    
    def compare_headers(self, df1, df2, output_wb):
        """Compare and highlight header differences"""
        headers1 = set(df1.columns)
        headers2 = set(df2.columns)
        common = headers1 & headers2
        unique1 = headers1 - headers2
        unique2 = headers2 - headers1
        
        # Create header comparison sheet
        ws = output_wb.create_sheet("Header Comparison")
        ws.append(["Header", "Status", "File 1 Presence", "File 2 Presence"])
        
        # Add common headers
        for header in sorted(common):
            ws.append([header, "Common", "✓", "✓"])
        
        # Add unique headers if option is enabled
        if self.highlight_missing.get():
            for header in sorted(unique1):
                ws.append([header, "Unique to File 1", "✓", ""])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
            
            for header in sorted(unique2):
                ws.append([header, "Unique to File 2", "", "✓"])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
        
        # Apply formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(bold=(cell.row == 1))
                cell.border = THIN_BORDER
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    
    def analyze_row_matches(self, df1, df2, output_wb, matched_count, unmatched1_count, unmatched2_count):
        """Analyze and highlight row matches between files"""
        # Create row matching sheet
        ws = output_wb.create_sheet("Row Matching Analysis")
        
        # Summary section
        ws.append(["Row Matching Summary"])
        ws.append(["", ""])
        ws.append(["Total Rows in File1", len(df1)])
        ws.append(["Total Rows in File2", len(df2)])
        ws.append(["Matched Rows", matched_count])
        ws.append(["Rows Only in File1", unmatched1_count])
        ws.append(["Rows Only in File2", unmatched2_count])
        ws.append([""])
        
        # Key generation method
        ws.append(["Key Generation Method:"])
        ws.append(["Automatically concatenated all non-date string columns"])
        ws.append([""])
        ws.append(["Note: Rows are considered matched if any row in File1 matches any row in File2"])
        
        # Apply styling to summary
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)
        
        for row in ws.iter_rows(min_row=3, max_row=7):
            for cell in row:
                cell.font = Font(bold=(cell.column == 1))
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    
    def compare_numeric_values(self, df1, df2, output_wb):
        """Create numerical comparison table for common numeric columns"""
        # Identify common numeric columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        num_cols = [col for col in common_cols 
                    if pd.api.types.is_numeric_dtype(df1[col]) and 
                    pd.api.types.is_numeric_dtype(df2[col])]
        
        if not num_cols:
            return
        
        # Create sheet
        ws = output_wb.create_sheet("Numeric Comparison")
        headers = ["Column", "Row", "File1 Value", "File2 Value", "Absolute Diff", "Relative Diff"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        # Compare values
        for col in num_cols:
            for i in range(min(len(df1), len(df2))):
                val1 = df1[col].iloc[i]
                val2 = df2[col].iloc[i]
                
                if pd.isna(val1) or pd.isna(val2) or val1 == val2:
                    continue
                    
                abs_diff = abs(val1 - val2)
                rel_diff = abs_diff / max(abs(val1), abs(val2)) if max(abs(val1), abs(val2)) != 0 else float('inf')
                
                ws.append([col, i+1, val1, val2, abs_diff, rel_diff])
                
                # Highlight significant differences (>10%)
                if rel_diff > 0.1:
                    for col_idx in range(1, 7):
                        cell = ws.cell(ws.max_row, col_idx)
                        cell.fill = NUM_DIFF_FILL
                
                # Apply borders
                for col_idx in range(1, 7):
                    cell = ws.cell(ws.max_row, col_idx)
                    cell.border = THIN_BORDER
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparator(root)
    root.mainloop()
