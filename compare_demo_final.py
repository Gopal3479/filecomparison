import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue
ROW_MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light Green
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")       # Light Gray

# Border style
THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

class ExcelComparator:
    def __init__(self, file1_path, file2_path, sheet1_name=None, sheet2_name=None, 
                 highlight_missing=True, highlight_cell_diffs=True, 
                 highlight_row_matches=True, create_num_table=True):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name
        self.highlight_missing = highlight_missing
        self.highlight_cell_diffs = highlight_cell_diffs
        self.highlight_row_matches = highlight_row_matches
        self.create_num_table = create_num_table
        
        if not self.sheet1_name:
            wb = load_workbook(file1_path, read_only=True)
            self.sheet1_name = wb.sheetnames[0]
            wb.close()
            
        if not self.sheet2_name:
            wb = load_workbook(file2_path, read_only=True)
            self.sheet2_name = wb.sheetnames[0]
            wb.close()
    
    def are_equal(self, a, b):
        """Check if two values are equal, handling NaN and date cases"""
        if pd.isna(a) and pd.isna(b):
            return True
        if pd.isna(a) or pd.isna(b):
            return False
        if isinstance(a, (datetime, pd.Timestamp)) and isinstance(b, (datetime, pd.Timestamp)):
            return a == b
        if isinstance(a, float) and isinstance(b, float):
            return round(a, 2) == round(b, 2)
        return str(a).strip() == str(b).strip()
    
    def compare_headers(self, df1, df2, output_wb):
        """Compare and highlight header differences"""
        headers1 = set(df1.columns)
        headers2 = set(df2.columns)
        common = headers1 & headers2
        unique1 = headers1 - headers2
        unique2 = headers2 - headers1
        
        ws = output_wb.create_sheet("Header Comparison")
        ws.append(["Header", "Status", "File 1 Presence", "File 2 Presence"])
        
        for header in sorted(common):
            ws.append([header, "Common", "✓", "✓"])
        
        if self.highlight_missing:
            for header in sorted(unique1):
                ws.append([header, "Unique to File 1", "✓", ""])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
            
            for header in sorted(unique2):
                ws.append([header, "Unique to File 2", "", "✓"])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(bold=(cell.row == 1))
                cell.border = THIN_BORDER
        
        for col_idx in range(1, 5):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def create_side_by_side_sheet(self, df1, df2, output_wb):
        """Create side-by-side comparison with proper row matching"""
        ws = output_wb.create_sheet("Side by Side Comparison")
        common_cols = list(set(df1.columns) & set(df2.columns))
        
        # Create header row
        header_row = list(df1.columns) + [" | "] + list(df2.columns) + ["Match Status"]
        ws.append(header_row)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        
        # Track row match status
        file1_match_status = []
        max_rows = max(len(df1), len(df2))
        
        for i in range(max_rows):
            row_match = True
            row_data = []
            
            # File1 data
            if i < len(df1):
                row_data.extend(df1.iloc[i].tolist())
            else:
                row_data.extend([""] * len(df1.columns))
                row_match = False
            
            # Separator
            row_data.append(" | ")
            
            # File2 data
            if i < len(df2):
                row_data.extend(df2.iloc[i].tolist())
            else:
                row_data.extend([""] * len(df2.columns))
                row_match = False
            
            # Compare values for common columns
            if i < len(df1) and i < len(df2):
                for col in common_cols:
                    val1 = df1.at[i, col]
                    val2 = df2.at[i, col]
                    if not self.are_equal(val1, val2):
                        row_match = False
                        break
            
            # Set match status
            status = "Matched" if row_match else "Not Matched"
            row_data.append(status)
            file1_match_status.append(status)
            
            ws.append(row_data)
        
        # Apply highlighting
        separator_col = len(df1.columns) + 1
        match_status_col = len(header_row)
        
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
            row_idx = i + 2  # Actual row index in worksheet
            
            # Apply border to all cells
            for cell in row:
                cell.border = THIN_BORDER
            
            # Highlight separator column
            ws.cell(row=row_idx, column=separator_col).fill = PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            )
            
            # Highlight row matches/mismatches
            if self.highlight_row_matches and i < len(file1_match_status):
                status = file1_match_status[i]
                fill = ROW_MATCH_FILL if status == "Matched" else ROW_MISSING_FILL
                
                # Apply to File1 section
                for col_idx in range(1, len(df1.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
                
                # Apply to File2 section
                for col_idx in range(len(df1.columns) + 2, len(df1.columns) + len(df2.columns) + 2):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
                
                # Apply to status column
                ws.cell(row=row_idx, column=match_status_col).fill = fill
            
            # Highlight cell differences
            if self.highlight_cell_diffs and i < min(len(df1), len(df2)):
                for col in common_cols:
                    val1 = df1.at[i, col]
                    val2 = df2.at[i, col]
                    
                    if not self.are_equal(val1, val2):
                        # Find column positions
                        col_idx1 = list(df1.columns).index(col) + 1
                        col_idx2 = list(df2.columns).index(col) + len(df1.columns) + 2
                        
                        # Apply highlighting
                        ws.cell(row=row_idx, column=col_idx1).fill = CELL_DIFF_FILL
                        ws.cell(row=row_idx, column=col_idx2).fill = CELL_DIFF_FILL
        
        # Auto-size columns
        for col_idx in range(1, len(header_row) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_idx]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Calculate match statistics
        matched_count = file1_match_status.count("Matched")
        unmatched1_count = len(df1) - matched_count
        unmatched2_count = len(df2) - matched_count
        
        return matched_count, unmatched1_count, unmatched2_count
    
    def analyze_row_matches(self, df1, df2, output_wb, matched_count, unmatched1_count, unmatched2_count):
        """Analyze row matches with accurate description"""
        ws = output_wb.create_sheet("Row Matching Analysis")
        
        # Summary section
        ws.append(["Row Matching Summary"])
        ws.append(["", ""])
        ws.append(["Total Rows in File1", len(df1)])
        ws.append(["Total Rows in File2", len(df2)])
        ws.append(["Matched Rows", matched_count])
        ws.append(["Unmatched Rows in File1", unmatched1_count])
        ws.append(["Unmatched Rows in File2", unmatched2_count])
        ws.append([""])
        
        # Matching method description
        ws.append(["Matching Method:"])
        ws.append(["Rows compared by index position"])
        ws.append(["Values considered equal if:"])
        ws.append(["  - Both are NaN/missing"])
        ws.append(["  - Numeric values rounded to 2 decimals match"])
        ws.append(["  - Strings match after stripping whitespace"])
        ws.append(["  - Dates match exactly"])
        
        # Apply styling
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)
        
        for row in ws.iter_rows(min_row=3, max_row=9):
            for cell in row:
                if cell.column == 1:
                    cell.font = Font(bold=True)
        
        # Auto-size columns
        for col_idx in range(1, 3):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def compare_numeric_values(self, df1, df2, output_wb):
        """Create numerical comparison table"""
        common_cols = list(set(df1.columns) & set(df2.columns))
        num_cols = [col for col in common_cols 
                    if pd.api.types.is_numeric_dtype(df1[col]) and 
                    pd.api.types.is_numeric_dtype(df2[col])]
        
        if not num_cols:
            return
        
        ws = output_wb.create_sheet("Numeric Comparison")
        headers = ["Column", "Row", "File1 Value", "File2 Value", "Absolute Diff", "Relative Diff"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        # Compare values
        for col in num_cols:
            for i in range(min(len(df1), len(df2))):
                val1 = df1[col].iloc[i]
                val2 = df2[col].iloc[i]
                
                if pd.isna(val1) or pd.isna(val2) or val1 == val2:
                    continue
                    
                abs_diff = abs(val1 - val2)
                base_val = max(abs(val1), abs(val2))
                rel_diff = abs_diff / base_val if base_val != 0 else float('inf')
                
                row_data = [col, i+1, val1, val2, abs_diff, rel_diff]
                ws.append(row_data)
                
                # Highlight significant differences
                if rel_diff > 0.1:
                    for col_idx in range(1, 7):
                        cell = ws.cell(ws.max_row, col_idx)
                        cell.fill = NUM_DIFF_FILL
                
                # Apply borders
                for col_idx in range(1, 7):
                    cell = ws.cell(ws.max_row, col_idx)
                    cell.border = THIN_BORDER
        
        # Auto-size columns
        for col_idx in range(1, 7):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    
    def compare(self, output_file=None):
        try:
            # Read data
            df1 = pd.read_excel(self.file1_path, sheet_name=self.sheet1_name)
            df2 = pd.read_excel(self.file2_path, sheet_name=self.sheet2_name)
            
            # Create comparison workbook
            output_wb = Workbook()
            output_wb.remove(output_wb.active)
            
            # 1. Compare headers
            self.compare_headers(df1, df2, output_wb)
            
            # 2. Create side-by-side comparison
            matched_count, unmatched1_count, unmatched2_count = self.create_side_by_side_sheet(df1, df2, output_wb)
            
            # 3. Row matching analysis
            self.analyze_row_matches(df1, df2, output_wb, matched_count, unmatched1_count, unmatched2_count)
            
            # 4. Numerical differences
            if self.create_num_table:
                self.compare_numeric_values(df1, df2, output_wb)
            
            if output_file:
                output_wb.save(output_file)
                return output_file
            return output_wb
                
        except Exception as e:
            raise Exception(f"Comparison error: {str(e)}")

# Example usage
if __name__ == "__main__":
    comparator = ExcelComparator(
        file1_path="file1.xlsx",
        file2_path="file2.xlsx",
        sheet1_name="Sheet1",
        sheet2_name="Sheet1",
        highlight_missing=True,
        highlight_cell_diffs=True,
        highlight_row_matches=True,
        create_num_table=True
    )
    
    try:
        result_path = comparator.compare(output_file="comparison_results.xlsx")
        print(f"Comparison saved to: {result_path}")
    except Exception as e:
        print(f"Error: {e}")
