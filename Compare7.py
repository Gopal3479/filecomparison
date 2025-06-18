import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import os
from datetime import datetime

# Define highlighting styles
HEADER_DIFF_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
CELL_DIFF_FILL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Tomato
NUM_DIFF_FILL = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")     # Light Sky Blue
ROW_MATCH_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light Green
ROW_MISSING_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
CONCAT_KEY_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # Light Pink

class ExcelComparator:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Excel Data Comparison Tool")
        self.root.geometry("1100x800")
        self.root.configure(bg="#f0f2f5")
        
        # Initialize variables
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.sheet1_name = tk.StringVar()
        self.sheet2_name = tk.StringVar()
        self.status = tk.StringVar(value="Ready to compare files")
        self.concat_columns = []
        self.df1 = None
        self.df2 = None
        
        # Create UI
        self.create_widgets()
        
    def create_widgets(self):
        # Header frame
        header_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        header_frame.pack(fill="x", side="top")
        
        header_label = tk.Label(
            header_frame, 
            text="Advanced Excel Data Comparison Tool", 
            font=("Arial", 20, "bold"), 
            fg="white", 
            bg="#2c3e50"
        )
        header_label.pack(pady=20)
        
        # Main content frame
        main_frame = tk.Frame(self.root, bg="#f0f2f5")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # File selection section
        file_frame = tk.LabelFrame(
            main_frame, 
            text="File Selection", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        file_frame.pack(fill="x", pady=(0, 15))
        
        # File 1
        file1_frame = tk.Frame(file_frame, bg="#f0f2f5")
        file1_frame.pack(fill="x", pady=5)
        
        tk.Label(
            file1_frame, 
            text="File 1:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=10, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            file1_frame, 
            textvariable=self.file1_path, 
            width=50, 
            state="readonly",
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        tk.Button(
            file1_frame, 
            text="Browse", 
            command=lambda: self.select_file(1), 
            bg="#3498db", 
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left")
        
        # File 2
        file2_frame = tk.Frame(file_frame, bg="#f0f2f5")
        file2_frame.pack(fill="x", pady=5)
        
        tk.Label(
            file2_frame, 
            text="File 2:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=10, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            file2_frame, 
            textvariable=self.file2_path, 
            width=50, 
            state="readonly",
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        tk.Button(
            file2_frame, 
            text="Browse", 
            command=lambda: self.select_file(2), 
            bg="#3498db", 
            fg="white",
            font=("Arial", 10, "bold")
        ).pack(side="left")
        
        # Sheet selection section
        sheet_frame = tk.LabelFrame(
            main_frame, 
            text="Sheet Selection", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        sheet_frame.pack(fill="x", pady=(0, 15))
        
        # Sheet 1
        sheet1_frame = tk.Frame(sheet_frame, bg="#f0f2f5")
        sheet1_frame.pack(fill="x", pady=5)
        
        tk.Label(
            sheet1_frame, 
            text="File 1 Sheet:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=15, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            sheet1_frame, 
            textvariable=self.sheet1_name, 
            width=30, 
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        # Sheet 2
        sheet2_frame = tk.Frame(sheet_frame, bg="#f0f2f5")
        sheet2_frame.pack(fill="x", pady=5)
        
        tk.Label(
            sheet2_frame, 
            text="File 2 Sheet:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=15, 
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            sheet2_frame, 
            textvariable=self.sheet2_name, 
            width=30, 
            font=("Arial", 10)
        ).pack(side="left", padx=5, fill="x", expand=True)
        
        # Concatenation columns selection
        concat_frame = tk.LabelFrame(
            main_frame, 
            text="String Columns for Row Matching", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        concat_frame.pack(fill="x", pady=(0, 15))
        
        concat_top_frame = tk.Frame(concat_frame, bg="#f0f2f5")
        concat_top_frame.pack(fill="x", pady=5)
        
        tk.Label(
            concat_top_frame, 
            text="Select columns for concatenation key:", 
            font=("Arial", 10), 
            bg="#f0f2f5", 
            width=25, 
            anchor="w"
        ).pack(side="left", padx=5)
        
        self.refresh_concat_btn = tk.Button(
            concat_top_frame, 
            text="Refresh Columns", 
            command=self.refresh_concat_columns,
            bg="#9b59b6", 
            fg="white",
            font=("Arial", 9, "bold")
        )
        self.refresh_concat_btn.pack(side="right", padx=5)
        
        self.concat_listbox = tk.Listbox(
            concat_frame, 
            selectmode=tk.MULTIPLE, 
            width=50, 
            height=6,
            font=("Arial", 9)
        )
        self.concat_listbox.pack(fill="x", padx=5, pady=5)
        
        # Options section
        options_frame = tk.LabelFrame(
            main_frame, 
            text="Comparison Options", 
            font=("Arial", 12, "bold"), 
            bg="#f0f2f5", 
            padx=10, 
            pady=10
        )
        options_frame.pack(fill="x", pady=(0, 15))
        
        self.highlight_missing = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight missing columns", 
            variable=self.highlight_missing, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.highlight_cell_diffs = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight cell differences", 
            variable=self.highlight_cell_diffs, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.highlight_row_matches = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Highlight row matches/mismatches", 
            variable=self.highlight_row_matches, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        self.create_num_table = tk.BooleanVar(value=True)
        tk.Checkbutton(
            options_frame, 
            text="Create numerical differences table", 
            variable=self.create_num_table, 
            bg="#f0f2f5", 
            font=("Arial", 10)
        ).pack(anchor="w", pady=3)
        
        # Action buttons
        button_frame = tk.Frame(main_frame, bg="#f0f2f5")
        button_frame.pack(fill="x", pady=20)
        
        compare_btn = tk.Button(
            button_frame, 
            text="Compare Excel Files", 
            command=self.compare_files, 
            bg="#27ae60", 
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=20
        )
        compare_btn.pack(pady=10)
        
        # Status bar
        status_frame = tk.Frame(self.root, bg="#e0e0e0", height=30)
        status_frame.pack(fill="x", side="bottom")
        
        tk.Label(
            status_frame, 
            textvariable=self.status, 
            font=("Arial", 10), 
            bg="#e0e0e0", 
            anchor="w"
        ).pack(side="left", padx=10)
        
        # Legend
        legend_frame = tk.Frame(main_frame, bg="#f0f2f5")
        legend_frame.pack(fill="x", pady=10)
        
        tk.Label(
            legend_frame, 
            text="Highlight Legend:", 
            font=("Arial", 10, "bold"), 
            bg="#f0f2f5"
        ).pack(anchor="w")
        
        legend_inner = tk.Frame(legend_frame, bg="#f0f2f5")
        legend_inner.pack(fill="x", pady=5)
        
        # Header diff legend
        header_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        header_legend.pack(side="left", padx=10)
        tk.Label(
            header_legend, 
            text="    ", 
            bg="#FFD700", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            header_legend, 
            text="Column Header Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Cell diff legend
        cell_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        cell_legend.pack(side="left", padx=10)
        tk.Label(
            cell_legend, 
            text="    ", 
            bg="#FF6347", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            cell_legend, 
            text="Cell Value Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Row match legend
        row_match_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        row_match_legend.pack(side="left", padx=10)
        tk.Label(
            row_match_legend, 
            text="    ", 
            bg="#90EE90", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            row_match_legend, 
            text="Row Matches", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Row missing legend
        row_missing_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        row_missing_legend.pack(side="left", padx=10)
        tk.Label(
            row_missing_legend, 
            text="    ", 
            bg="#D3D3D3", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            row_missing_legend, 
            text="Missing Rows", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Concat key legend
        concat_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        concat_legend.pack(side="left", padx=10)
        tk.Label(
            concat_legend, 
            text="    ", 
            bg="#FFB6C1", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            concat_legend, 
            text="Concatenation Key", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
        
        # Num diff legend
        num_legend = tk.Frame(legend_inner, bg="#f0f2f5")
        num_legend.pack(side="left", padx=10)
        tk.Label(
            num_legend, 
            text="    ", 
            bg="#87CEFA", 
            width=3, 
            height=1
        ).pack(side="left")
        tk.Label(
            num_legend, 
            text="Numerical Differences", 
            font=("Arial", 9), 
            bg="#f0f2f5"
        ).pack(side="left", padx=5)
    
    def select_file(self, file_num):
        file_path = filedialog.askopenfilename(
            title=f"Select Excel File {file_num}",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            if file_num == 1:
                self.file1_path.set(file_path)
                try:
                    wb = load_workbook(file_path, read_only=True)
                    self.sheet1_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet1_name.set("Sheet1")
            else:
                self.file2_path.set(file_path)
                try:
                    wb = load_workbook(file_path, read_only=True)
                    self.sheet2_name.set(wb.sheetnames[0])
                    wb.close()
                except:
                    self.sheet2_name.set("Sheet1")
    
    def refresh_concat_columns(self):
        """Refresh available string columns for concatenation"""
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()
        sheet1 = self.sheet1_name.get()
        sheet2 = self.sheet2_name.get()
        
        if not file1 or not file2:
            messagebox.showerror("Error", "Please select both Excel files first")
            return
        
        try:
            # Read data
            df1 = pd.read_excel(file1, sheet_name=sheet1)
            df2 = pd.read_excel(file2, sheet_name=sheet2)
            
            # Find common columns that are string type and not dates
            common_cols = list(set(df1.columns) & set(df2.columns))
            string_cols = []
            
            for col in common_cols:
                # Check if column is string type and not date
                if (pd.api.types.is_string_dtype(df1[col]) and 
                    pd.api.types.is_string_dtype(df2[col]) and
                    not pd.api.types.is_datetime64_any_dtype(df1[col]) and
                    not pd.api.types.is_datetime64_any_dtype(df2[col])):
                    string_cols.append(col)
            
            # Update listbox
            self.concat_listbox.delete(0, tk.END)
            for col in string_cols:
                self.concat_listbox.insert(tk.END, col)
                
            if not string_cols:
                self.concat_listbox.insert(tk.END, "No suitable string columns found")
                
            self.status.set(f"Found {len(string_cols)} string columns for concatenation")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not read files:\n{str(e)}")
            self.status.set("Error reading files")
    
    def are_equal(self, a, b):
        """Check if two values are equal, handling NaN cases"""
        if pd.isna(a) and pd.isna(b):
            return True
        if pd.isna(a) or pd.isna(b):
            return False
        return a == b
    
    def is_date(self, value):
        """Check if a value is a date"""
        return isinstance(value, (datetime, pd.Timestamp))
    
    def compare_files(self):
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()
        sheet1 = self.sheet1_name.get()
        sheet2 = self.sheet2_name.get()
        
        if not file1 or not file2:
            messagebox.showerror("Error", "Please select both Excel files")
            return
        
        # Get selected concatenation columns
        self.concat_columns = [self.concat_listbox.get(i) 
                              for i in self.concat_listbox.curselection()]
        
        if not self.concat_columns:
            messagebox.showwarning("Warning", 
                                  "No columns selected for concatenation key. "
                                  "Row matching will not be performed.")
        
        self.status.set("Reading files...")
        self.root.update()
        
        try:
            # Read data
            self.df1 = pd.read_excel(file1, sheet_name=sheet1)
            self.df2 = pd.read_excel(file2, sheet_name=sheet2)
            
            # Pad dataframes to same length
            max_len = max(len(self.df1), len(self.df2))
            df1_padded = self.df1.reindex(range(max_len))
            df2_padded = self.df2.reindex(range(max_len))
            
            # Create comparison workbook
            output_wb = Workbook()
            output_wb.remove(output_wb.active)
            
            # 1. Compare headers
            self.compare_headers(df1_padded, df2_padded, output_wb)
            
            # 2. Compare row data with row matching
            self.compare_row_data(df1_padded, df2_padded, output_wb)
            
            # 3. Row matching analysis
            if self.concat_columns:
                self.analyze_row_matches(df1_padded, df2_padded, output_wb)
            
            # 4. Numerical differences
            if self.create_num_table.get():
                self.compare_numeric_values(df1_padded, df2_padded, output_wb)
            
            # Save results
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Comparison Results"
            )
            
            if output_file:
                output_wb.save(output_file)
                self.status.set(f"Comparison saved to: {output_file}")
                messagebox.showinfo("Success", f"Comparison saved successfully!\n{output_file}")
            else:
                self.status.set("Comparison canceled")
                
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.status.set("Error occurred - see details in message")
    
    def compare_headers(self, df1, df2, output_wb):
        """Compare and highlight header differences"""
        headers1 = set(df1.columns)
        headers2 = set(df2.columns)
        common = headers1 & headers2
        unique1 = headers1 - headers2
        unique2 = headers2 - headers1
        
        # Create header comparison sheet
        ws = output_wb.create_sheet("Header Comparison")
        ws.append(["Header", "Status", "File 1 Presence", "File 2 Presence"])
        
        # Add common headers
        for header in sorted(common):
            ws.append([header, "Common", "✓", "✓"])
        
        # Add unique headers if option is enabled
        if self.highlight_missing.get():
            for header in sorted(unique1):
                ws.append([header, "Unique to File 1", "✓", ""])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
            
            for header in sorted(unique2):
                ws.append([header, "Unique to File 2", "", "✓"])
                ws.cell(ws.max_row, 1).fill = HEADER_DIFF_FILL
        
        # Apply formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    def compare_row_data(self, df1, df2, output_wb):
        """Compare row-wise data and highlight differences"""
        # Get common columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        
        # Create comparison sheets
        ws1 = output_wb.create_sheet("File1 Data")
        ws2 = output_wb.create_sheet("File2 Data")
        
        # Write headers
        for col_idx, header in enumerate(df1.columns, 1):
            ws1.cell(1, col_idx, header)
            ws1.cell(1, col_idx).font = Font(bold=True)
        
        for col_idx, header in enumerate(df2.columns, 1):
            ws2.cell(1, col_idx, header)
            ws2.cell(1, col_idx).font = Font(bold=True)
        
        # Create column position mappings
        col_pos1 = {col: idx for idx, col in enumerate(df1.columns, 1)}
        col_pos2 = {col: idx for idx, col in enumerate(df2.columns, 1)}
        
        # Create concatenation keys if columns are selected
        concat_keys1 = {}
        concat_keys2 = {}
        
        if self.concat_columns:
            # Create concatenated keys
            for idx, row in df1.iterrows():
                key_parts = []
                for col in self.concat_columns:
                    if col in df1.columns and not self.is_date(row[col]):
                        key_parts.append(str(row[col]))
                concat_keys1[idx] = "_".join(key_parts) if key_parts else None
            
            for idx, row in df2.iterrows():
                key_parts = []
                for col in self.concat_columns:
                    if col in df2.columns and not self.is_date(row[col]):
                        key_parts.append(str(row[col]))
                concat_keys2[idx] = "_".join(key_parts) if key_parts else None
            
            # Create sets of keys for matching
            keys1_set = set(concat_keys1.values())
            keys2_set = set(concat_keys2.values())
        
        # Write data and highlight differences
        for row_idx in range(len(df1)):
            # Write full row for File1
            for col in df1.columns:
                val = df1[col].iloc[row_idx]
                ws1.cell(row_idx+2, col_pos1[col], val)
            
            # Write full row for File2
            for col in df2.columns:
                val = df2[col].iloc[row_idx]
                ws2.cell(row_idx+2, col_pos2[col], val)
            
            # Highlight row matching status if enabled
            if self.highlight_row_matches.get() and self.concat_columns:
                key1 = concat_keys1.get(row_idx)
                key2 = concat_keys2.get(row_idx)
                
                # Highlight if row exists in both files
                if key1 and key1 in keys2_set:
                    for col in df1.columns:
                        ws1.cell(row_idx+2, col_pos1[col]).fill = ROW_MATCH_FILL
                
                # Highlight if row exists in both files
                if key2 and key2 in keys1_set:
                    for col in df2.columns:
                        ws2.cell(row_idx+2, col_pos2[col]).fill = ROW_MATCH_FILL
            
            # Highlight differences in common columns
            if self.highlight_cell_diffs.get():
                for col in common_cols:
                    val1 = df1[col].iloc[row_idx]
                    val2 = df2[col].iloc[row_idx]
                    
                    if not self.are_equal(val1, val2):
                        ws1.cell(row_idx+2, col_pos1[col]).fill = CELL_DIFF_FILL
                        ws2.cell(row_idx+2, col_pos2[col]).fill = CELL_DIFF_FILL
    
    def analyze_row_matches(self, df1, df2, output_wb):
        """Analyze and highlight row matches between files using concatenated keys"""
        if not self.concat_columns:
            return
        
        # Create row matching sheet
        ws = output_wb.create_sheet("Row Matching Analysis")
        ws.append(["Concatenation Key", "Status", "File 1 Row", "File 2 Row", "Key Components"])
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Create concatenated keys
        concat_keys1 = {}
        concat_keys2 = {}
        
        for idx, row in df1.iterrows():
            key_parts = []
            for col in self.concat_columns:
                if col in df1.columns and not self.is_date(row[col]):
                    key_parts.append(str(row[col]))
            concat_keys1[idx] = "_".join(key_parts) if key_parts else None
        
        for idx, row in df2.iterrows():
            key_parts = []
            for col in self.concat_columns:
                if col in df2.columns and not self.is_date(row[col]):
                    key_parts.append(str(row[col]))
            concat_keys2[idx] = "_".join(key_parts) if key_parts else None
        
        # Get unique keys from both files
        keys1 = set(concat_keys1.values())
        keys2 = set(concat_keys2.values())
        
        # Remove None values
        keys1.discard(None)
        keys2.discard(None)
        
        # Find common keys
        common_keys = keys1 & keys2
        
        # Find keys only in file1
        only_in_file1 = keys1 - keys2
        
        # Find keys only in file2
        only_in_file2 = keys2 - keys1
        
        # Add common rows
        for key in sorted(common_keys):
            row1 = [idx for idx, k in concat_keys1.items() if k == key][0] + 1
            row2 = [idx for idx, k in concat_keys2.items() if k == key][0] + 1
            key_components = ", ".join(self.concat_columns)
            ws.append([key, "Present in both files", row1, row2, key_components])
            ws.cell(ws.max_row, 1).fill = CONCAT_KEY_FILL
        
        # Add rows only in file1
        for key in sorted(only_in_file1):
            row1 = [idx for idx, k in concat_keys1.items() if k == key][0] + 1
            key_components = ", ".join(self.concat_columns)
            ws.append([key, "Only in File 1", row1, "N/A", key_components])
            ws.cell(ws.max_row, 1).fill = ROW_MISSING_FILL
        
        # Add rows only in file2
        for key in sorted(only_in_file2):
            row2 = [idx for idx, k in concat_keys2.items() if k == key][0] + 1
            key_components = ", ".join(self.concat_columns)
            ws.append([key, "Only in File 2", "N/A", row2, key_components])
            ws.cell(ws.max_row, 1).fill = ROW_MISSING_FILL)
    
    def compare_numeric_values(self, df1, df2, output_wb):
        """Create numerical comparison table for common numeric columns"""
        # Identify common numeric columns
        common_cols = list(set(df1.columns) & set(df2.columns))
        num_cols = [col for col in common_cols 
                    if pd.api.types.is_numeric_dtype(df1[col]) and 
                    pd.api.types.is_numeric_dtype(df2[col])]
        
        if not num_cols:
            return
        
        # Create sheet
        ws = output_wb.create_sheet("Numeric Comparison")
        headers = ["Column", "Row", "File1 Value", "File2 Value", "Absolute Diff", "Relative Diff"]
        ws.append(headers)
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Compare values
        for col in num_cols:
            for i in range(min(len(df1), len(df2))):
                val1 = df1[col].iloc[i]
                val2 = df2[col].iloc[i]
                
                if pd.isna(val1) or pd.isna(val2) or val1 == val2:
                    continue
                    
                abs_diff = abs(val1 - val2)
                rel_diff = abs_diff / max(abs(val1), abs(val2)) if max(abs(val1), abs(val2)) != 0 else float('inf')
                
                ws.append([col, i+1, val1, val2, abs_diff, rel_diff])
                
                # Highlight significant differences (>10%)
                if rel_diff > 0.1:
                    for col_idx in range(1, 7):
                        ws.cell(ws.max_row, col_idx).fill = NUM_DIFF_FILL

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparator(root)
    root.mainloop()
